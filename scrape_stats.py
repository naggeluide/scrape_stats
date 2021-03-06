#!/usr/bin/env python

import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import datetime
from openpyxl import load_workbook
import numpy as np
import os

#https://github.com/alexwlchan/ao3/blob/936cce2684e4ea506e3d6eb8a765ca2c5c8c31ef/src/ao3/users.py#L26-L41

#made it a chronjob here: https://stackabuse.com/scheduling-jobs-with-python-crontab/

def read_from_session(login_data, baseurl='https://archiveofourown.org/'):
    login_url = baseurl+'users/login'
    stats_url=baseurl+'users/'+login_data['username']+'/stats'
    works_url=baseurl+'users/'+login_data['username']+'/works'
    short_wurl='/users/'+login_data['username']+'/works?page='

    #make the soup
    with requests.Session() as s:
        req = s.get(baseurl,allow_redirects=True)
        soup0 = BeautifulSoup(req.text, features='html.parser')
        authenticity_token = soup0.find('input', {'name': 'authenticity_token'})['value']
        time.sleep(2)
        req = s.post(login_url, data={
            'authenticity_token': authenticity_token,
            'user[login]': login_data['username'],
            'user[password]': login_data['password'],})
        time.sleep(2)
        #print req.text
        stats_page=s.get(stats_url)
        soup = BeautifulSoup(stats_page.text, 'html.parser')
        #print soup.prettify()
        #get works pages for public bookmarks, total comments and collection stats -- update this so it takes n pages of works
        works_page1=s.get(works_url+'?page=1')
        wsoup1 = BeautifulSoup(works_page1.text, 'html.parser')
        #all stats from works page
        works1 = wsoup1.find_all('li', attrs = {'class':'own work blurb group'})
        #how many pages?
        #ex: a href="/users/naggeluide/works?page=2"
        pa = wsoup1.find('ol', attrs = {'class':'pagination actions'})
        patext=pa.text
        pp=patext.find('Previous')+9
        pn=patext.find('Next')-1
        npages=patext[pp:pn]
        nlist=npages.split(' ')
        for n in nlist:
            works_pagen=s.get(works_url+'?page='+n)
            wsoupn = BeautifulSoup(works_pagen.text, 'html.parser')
            worksn = wsoupn.find_all('li', attrs = {'class':'own work blurb group'})
            for w in worksn:
                works1.append(w)
    tstamp=datetime.datetime.now()
    #get interesting stuff
    all_stats=soup.find_all('dl', attrs = {'class':'stats'}) #use to find those without subs...
    titl=soup.find_all('a', {'href': lambda x : x.startswith('/works/')})
    titles=[t.text for t in titl]
    for i in range(0,4):
       titles.pop(0)
    return tstamp,soup,titles,all_stats,works1

def total_stats(tstamp,soup):
    #cover sheet - total stats. make dataframe
    us=int(soup.find('dd', attrs = {'class':'user subscriptions'}).text)
    tk=int(soup.find('dd', attrs = {'class':'kudos'}).text)
    tct=int(soup.find('dd', attrs = {'class':'comment thread count'}).text)
    tbm=int(soup.find('dd', attrs = {'class':'bookmarks'}).text)
    ts=int(soup.find_all('dd', attrs = {'class':'subscriptions'})[1].text)
    twc=int(soup.find('dd', attrs = {'class':'words'}).text)
    th=int(soup.find('dd', attrs = {'class':'hits'}).text)
    #make into a dataframe
    tstats=pd.DataFrame({"User Subs":us,"Total Kudos":tk,"Total Comment Threads":tct,"Total Bookmarks":tbm,"Total Subscriptions":ts,"Total Word Count": twc,"Total Hits":th},index=[tstamp])
    return tstats

def get_all_stats(soup,all_stats):
    wc,kk,cc,bm,ss,hh=[],[],[],[],[],[]
    wc=[t.text for t in soup.find_all('span', attrs = {'class':'words'})]
    wcp=[int(w[1:-6]) for w in wc]
    kk=[int(t.text) for t in soup.find_all('dd', attrs = {'class':'kudos'})] #+total
    cc=[int(t.text) for t in soup.find_all('dd', attrs = {'class':'comments'})]
    bm=[int(t.text) for t in soup.find_all('dd', attrs = {'class':'bookmarks'})] #+total
    ss=[int(t.text) for t in soup.find_all('dd', attrs = {'class':'subscriptions'})] #this is + user, total, - those with none
    idx=[]
    for i,a in enumerate(all_stats):
        if 'Subscriptions' not in a.text:
            idx.append(i)
    ss=ss[2:] #get rid of user and total
    for i in idx:
        ss.insert(i,0)
    hh=[int(t.text) for t in soup.find_all('dd', attrs = {'class':'hits'})] #+total
    return np.array(zip(wcp,kk[1:],cc,bm[1:],ss,hh[1:]))

def get_story_stats_ext(title,works):
    ''' get external stats for a single work, ie public bookmarks and collections'''
    #get fandom too, stick that in first row
    for w in works:
        bb=w.text
        bms=0
        ncs=0
        if bb.find(title) !=-1: #in the first 100 chars just in case it's a common word
            if bb.find('Bookmarks') !=-1:
                bms=int(bb[bb.find('Bookmarks')+11:bb.find('Hits')-1])
            if bb.find('Collections') != -1:
                ncs=int(bb[bb.find('Collections')+13:bb.find('Comments')-1])
            break
    return bms,ncs,w

def single_work_df(tstamp,astats,bms,ncs):
    '''combine everything into a single row dataframe'''
    sdict={}
    sdict['Word Count']=astats[0]
    sdict['Kudos']=astats[1]
    sdict['Comment Threads']=astats[2]
    sdict['Bookmarks']=astats[3]
    sdict['Subscriptions']=astats[4]
    sdict['Hits']=astats[5]
    sdict['Public Bookmarks']=bms
    sdict['Collections']=ncs
    work_df=pd.DataFrame(sdict,index=[tstamp])
    return work_df

def meta_to_df(meta):
    '''put metadata -- fandom, tags, etc -- into df for writing'''
    #<h5 class="fandoms heading">
    fandom_head=meta.find('h5',attrs={'class':'fandoms heading'})
    fandom=pd.Series(fandom_head.find('a').text)
    #<ul class="required-tags">
    rtags=meta.find('ul',attrs={'class':'required-tags'}).text.split('\n')
    rating=pd.Series(rtags[1])
    warnings=pd.Series(rtags[2:-3])
    cat=pd.Series(rtags[-3])
    status=pd.Series(rtags[-2])
    #<li class="relationships">
    reltags=meta.find_all('li',attrs={'class':'relationships'})
    if len(reltags) !=0:
        reltext=pd.Series([r.text for r in reltags])
    else:
        reltext=pd.Series(None)
    #<li class="characters">
    charac=meta.find_all('li',attrs={'class':'characters'})
    chartext=pd.Series([r.text for r in charac])
    #<li class="freeforms">
    freef=meta.find_all('li',attrs={'class':'freeforms'})
    if len(freef) !=0:
        freetext=pd.Series([r.text for r in freef])
    else:
        freetext=pd.Series(None)
    #<dd class="language">
    lang=pd.Series(meta.find('dd',attrs={'class':'language'}).text)

    #now put in a dataframe
    mdict={'Fandom':fandom,'Rating':rating,'Warnings':warnings,'Category':cat,'Status':status,'Relationships':reltext,'Characters':chartext,'Additional Tags':freetext,'Language':lang}
    meta_df=pd.DataFrame(mdict)
    return meta_df


def fix_title(t):
    fchars=[':','/','\\','?','*','[',']'] #or any other character forbidden to Excel... : \ / ? * [ ]
    for f in fchars:
        if f in t:
            t=t.replace(f,'-')
    if len(t) >31:
        t=t[:31]
    return t

def write_to_sheet(t,df,xls_file,meta=False): #append...check that this works
    #https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    book=load_workbook(xls_file)
    writer= pd.ExcelWriter(xls_file, engine = 'openpyxl')# as writer:
    writer.book=book
    t=fix_title(t)
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    if t not in writer.sheets.keys(): #create page
        book.create_sheet(t)
        df.to_excel(writer,sheet_name=t)
    elif not meta:
        try: #true for works
            startrow= df['Hits'].count()+3 #one for the header, one for starts on 1. this discounts the NaNs that may occur if meta columns are longer than stats columns
        except KeyError: #Totals page
            startrow = len(writer.book[t].columns[0])#writer.book[t].max_row
        df.to_excel(writer,sheet_name=t,startrow=startrow,header=False)
    if meta: #add meta columns...
        startcol=10 #K
        df.to_excel(writer,sheet_name=t,startcol=startcol,header=True)
    writer.save()
    writer.close()



if __name__ == "__main__":
    login_data = {
    'username': 'your_username NOT email',
    'password': 'your_password'
    }
    xls_file='/path/to/file/filename.xlsx'

    tstamp,soup,titles,all_stats,works1=read_from_session(login_data)
    tstats=total_stats(tstamp,soup)
    astats=get_all_stats(soup,all_stats)

    if os.path.isfile(xls_file)==False: #write initial file
        writer= pd.ExcelWriter(xls_file, engine = 'xlsxwriter')# as writer:
        tstats.to_excel(writer, sheet_name='Totals')
        for i,t in enumerate(set(titles)):
            bms,ncs,meta=get_story_stats_ext('\n'+t+'\n',works1)
            meta_df=meta_to_df(meta)
            work_df=single_work_df(tstamp,astats[i],bms,ncs)
            work_df.to_excel(writer,sheet_name=t)
            meta_df.to_excel(writer,sheet_name=t,meta=True)
        #writer.save()
        writer.close()

    else:

        write_to_sheet('Totals',tstats,xls_file)
        #need to stop crossovers from getting written twice...set() should do it
        #j/k that messes up because of all_stats...
        seen=[]
        for i,t in enumerate(titles):
            if t in seen:
                continue
            else:
                seen.append(t)
                bms,ncs,meta=get_story_stats_ext('\n'+t+'\n',works1)
                meta_df=meta_to_df(meta)
                work_df=single_work_df(tstamp,astats[i],bms,ncs)
                t=fix_title(t)
                write_to_sheet(t,work_df,xls_file)
                #did meta change? if so, write new meta. if not, ignore
                current_meta_df=pd.read_excel(xls_file, t,usecols=range(10,19))
                if not meta_df.equals(current_meta_df):
                    write_to_sheet(t,meta_df,xls_file,meta=True) #this just overwrites cuz I'm lazy
